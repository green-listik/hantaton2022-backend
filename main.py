import openpyxl
from openpyxl.styles import Alignment, Font


class Excel:
    def __init__(self, file_name: str):
        self.numb_in_sec = 9
        self.file_name = file_name
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def get_pattern(self, pattern_name: str):
        """takes an excel file pattern"""
        self.wb = openpyxl.load_workbook("./" + pattern_name)
        self.ws = self.wb.active
        return self.__file_save()

    def __file_save(self):
        """saves the file"""
        self.wb.save(self.file_name)

    def __create_cells(self, cell: str, value, **kwargs):
        """creating a cell according to the specified parameters"""
        self.ws[cell].value = value
        if "merge_cell" in kwargs.keys():
            self.ws.merge_cells(f"{cell}:{kwargs['merge_cell']}")
        if "font" in kwargs.keys():
            self.ws[cell].font = kwargs["font"]
        if "alignment" in kwargs.keys():
            self.ws[cell].alignment = kwargs["alignment"]
        return {cell, value}

    def create_section(self, dict_sec: list, font_=Font(size=10, bold=False)):
        """recording operations in an excel file"""
        for elem in dict_sec:
            string = elem["parameters"]
            self.__create_cells(f"A{self.numb_in_sec}", elem["id"], font=font_,
                                alignment=Alignment(horizontal="center", vertical="center"))
            self.__create_cells(f"B{self.numb_in_sec}", elem["name"], font=font_,
                                alignment=Alignment(vertical="center"))
            self.__create_cells(f"C{self.numb_in_sec}", string["planned time"], font=font_,
                                alignment=Alignment(vertical="center", horizontal="center"))

            if str(self.ws[f"D{self.numb_in_sec - 1}"].value).replace(".", "", 1).isdigit():
                self.__create_cells(f"D{self.numb_in_sec}", f"=C{self.numb_in_sec}/24+D{self.numb_in_sec - 1}",
                                    font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))
            else:
                self.__create_cells(f"D{self.numb_in_sec}", f"=C{self.numb_in_sec}/24", font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))

            self.__create_cells(f"E{self.numb_in_sec}", string["planned depth"], font=font_,
                                alignment=Alignment(vertical="center", horizontal="center"))
            self.__create_cells(f"F{self.numb_in_sec}", string["actual time"], font=font_,
                                alignment=Alignment(vertical="center", horizontal="center"))

            if str(self.ws[f"G{self.numb_in_sec - 1}"].value).replace(".", "", 1).isdigit():
                self.__create_cells(f"G{self.numb_in_sec}", f"=F{self.numb_in_sec}/24+G{self.numb_in_sec - 1}",
                                    font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))
            else:
                self.__create_cells(f"G{self.numb_in_sec}", f"=F{self.numb_in_sec}/24", font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))

            self.__create_cells(f"H{self.numb_in_sec}", string["actual depth"], font=font_,
                                alignment=Alignment(vertical="center", horizontal="center"))
            self.__create_cells(f"I{self.numb_in_sec}", f"=F{self.numb_in_sec}-C{self.numb_in_sec}", font=font_,
                                alignment=Alignment(vertical="center", horizontal="center"))

            if str(self.ws[f"J{self.numb_in_sec - 1}"].value).replace(".", "", 1).isdigit():
                self.__create_cells(f"J{self.numb_in_sec}", f"=I{self.numb_in_sec}/24+J{self.numb_in_sec - 1}",
                                    font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))
            else:
                self.__create_cells(f"J{self.numb_in_sec}", f"=I{self.numb_in_sec}/24", font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))

            if "actual NPT" in string.keys():
                self.__create_cells(f"K{self.numb_in_sec}", string["actual NPT"], font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))
            self.__create_cells(f"L{self.numb_in_sec}", f"=E{self.numb_in_sec}-K{self.numb_in_sec}", font=font_,
                                alignment=Alignment(vertical="center", horizontal="center"))

            if str(self.ws[f"M{self.numb_in_sec - 1}"].value).replace(".", "", 1).isdigit():
                self.__create_cells(f"M{self.numb_in_sec}", f"=K{self.numb_in_sec}/24+M{self.numb_in_sec - 1}",
                                    font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))
            else:
                self.__create_cells(f"M{self.numb_in_sec}", f"=K{self.numb_in_sec}/24", font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))

            if "Short details on NPT" in string.keys():
                self.__create_cells(f"N{self.numb_in_sec}", string["Short details on NPT"], font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))

            if "ILT or extra to mKPI (hrs)" in string.keys():
                self.__create_cells(f"O{self.numb_in_sec}", string["Short details on NPT"], font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))

            if str(self.ws[f"P{self.numb_in_sec - 1}"].value).replace(".", "", 1).isdigit():
                self.__create_cells(f"R{self.numb_in_sec}", f"=O{self.numb_in_sec}/24+P{self.numb_in_sec - 1}",
                                    font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))
            else:
                self.__create_cells(f"P{self.numb_in_sec}", f"=O{self.numb_in_sec}/24", font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))

            if str(self.ws[f"Q{self.numb_in_sec - 1}"].value).replace(".", "", 1).isdigit():
                self.__create_cells(f"Q{self.numb_in_sec}",
                                    f"=(F{self.numb_in_sec}-K{self.numb_in_sec}-O{self.numb_in_sec})/24+S{self.numb_in_sec - 1}",
                                    font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))
            else:
                self.__create_cells(f"Q{self.numb_in_sec}",
                                    f"=(F{self.numb_in_sec}-K{self.numb_in_sec}-O{self.numb_in_sec})/24", font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))

            if "Short details on ILT" in string.keys():
                self.__create_cells(f"R{self.numb_in_sec}", string["Short details on ILT"], font=font_,
                                    alignment=Alignment(vertical="center", horizontal="center"))

            self.numb_in_sec += 1
        return self.__file_save()


"""usage example"""
# a = Excel("111.xlsx")
# a.get_pattern("pattern.xlsx")
#
# a.create_section([
#     {"id": "010",
#      "name": "Preparations before skidding. Skid the rig for 10m from slot #11 to the slot #12. Preparations after skiding",
#      "parameters": {
#          "planned time": 2,
#          "planned depth": 0,
#          "hide column": -0.5,
#          "actual time": 2,
#          "actual depth": 34,
#          "actual NPT": 4,
#          "Short details on ILT": "Ream down in AС-6"
#      }},
#      {
#      "id": "100",
#      "name": "M/U steerable BHA#1.",
#     "parameters": {
#          "phase": "D12",
#          "name": "M/U steerable BHA#1.",
#          "planned time": 1.25,
#          "planned depth": 40,
#          "hide column": 2.34,
#          "actual time": 0.5,
#          "actual depth": 755,
#          "Short details on NPT": "NPT:SSC Mud Pump Failure 4 hrs"
#      }
#      }]
# )
